from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_CSV = ROOT / "data" / "processed" / "coffee_shop_sales_clean.csv"
PBI_DIR = ROOT / "powerbi" / "Retail Sales Revenue Dashboard"
PROJECT_NAME = "Retail Sales Revenue Dashboard"


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def m_csv_expression() -> list[str]:
    csv_path = str(DATA_CSV)
    return [
        "let",
        f'    Source = Csv.Document(File.Contents("{csv_path}"), [Delimiter=",", Columns=25, Encoding=65001, QuoteStyle=QuoteStyle.Csv]),',
        "    PromotedHeaders = Table.PromoteHeaders(Source, [PromoteAllScalars=true]),",
        "    ChangedTypes = Table.TransformColumnTypes(PromotedHeaders, {",
        '        {"OrderID", Int64.Type}, {"OrderDate", type date}, {"OrderTime", type time},',
        '        {"Quantity", Int64.Type}, {"StoreID", Int64.Type}, {"StoreLocation", type text},',
        '        {"City", type text}, {"Region", type text}, {"CustomerID", type text}, {"Segment", type text},',
        '        {"ProductID", Int64.Type}, {"ProductCategory", type text}, {"ProductType", type text}, {"ProductName", type text},',
        '        {"UnitPrice", Currency.Type}, {"Revenue", Currency.Type}, {"EstimatedCost", Currency.Type},',
        '        {"Profit", Currency.Type}, {"ProfitRate", Percentage.Type}, {"Year", Int64.Type},',
        '        {"MonthNo", Int64.Type}, {"MonthName", type text}, {"YearMonth", type text},',
        '        {"Weekday", type text}, {"Hour", Int64.Type}',
        "    })",
        "in",
        "    ChangedTypes",
    ]


def column(name: str, data_type: str, *, fmt: str | None = None, summarize_by: str = "none") -> dict:
    result = {
        "name": name,
        "dataType": data_type,
        "sourceColumn": name,
        "summarizeBy": summarize_by,
    }
    if fmt:
        result["formatString"] = fmt
    return result


def measure(name: str, expression: str, fmt: str | None = None) -> dict:
    result = {"name": name, "expression": expression}
    if fmt:
        result["formatString"] = fmt
    return result


def create_model_bim() -> None:
    sales_columns = [
        column("OrderID", "int64"),
        column("OrderDate", "dateTime", fmt="Short Date"),
        column("OrderTime", "dateTime"),
        column("Quantity", "int64", summarize_by="sum"),
        column("StoreID", "int64"),
        column("StoreLocation", "string"),
        column("City", "string"),
        column("Region", "string"),
        column("CustomerID", "string"),
        column("Segment", "string"),
        column("ProductID", "int64"),
        column("ProductCategory", "string"),
        column("ProductType", "string"),
        column("ProductName", "string"),
        column("UnitPrice", "decimal", fmt="$#,0.00"),
        column("Revenue", "decimal", fmt="$#,0.00", summarize_by="sum"),
        column("EstimatedCost", "decimal", fmt="$#,0.00", summarize_by="sum"),
        column("Profit", "decimal", fmt="$#,0.00", summarize_by="sum"),
        column("ProfitRate", "decimal", fmt="0.00%"),
        column("Year", "int64"),
        column("MonthNo", "int64"),
        column("MonthName", "string"),
        column("YearMonth", "string"),
        column("Weekday", "string"),
        column("Hour", "int64"),
    ]

    measures = [
        measure("Total Revenue", "SUM ( Sales[Revenue] )", "$#,0.00"),
        measure("Total Profit", "SUM ( Sales[Profit] )", "$#,0.00"),
        measure("Profit Margin", "DIVIDE ( [Total Profit], [Total Revenue] )", "0.00%"),
        measure("Total Orders", "DISTINCTCOUNT ( Sales[OrderID] )", "#,0"),
        measure("Total Quantity", "SUM ( Sales[Quantity] )", "#,0"),
        measure("Avg Order Value", "DIVIDE ( [Total Revenue], [Total Orders] )", "$#,0.00"),
        measure("Customer Count", "DISTINCTCOUNT ( Sales[CustomerID] )", "#,0"),
        measure("Avg Customer Spend", "DIVIDE ( [Total Revenue], [Customer Count] )", "$#,0.00"),
        measure(
            "Repeat Customer",
            "VAR CustomerOrderTable = SUMMARIZE ( Sales, Sales[CustomerID], \"OrderCount\", DISTINCTCOUNT ( Sales[OrderID] ) ) RETURN COUNTROWS ( FILTER ( CustomerOrderTable, [OrderCount] > 1 ) )",
            "#,0",
        ),
        measure("Repeat Customer Rate", "DIVIDE ( [Repeat Customer], [Customer Count] )", "0.00%"),
        measure("Profit Per Order", "DIVIDE ( [Total Profit], [Total Orders] )", "$#,0.00"),
        measure("Profit Per Unit", "DIVIDE ( [Total Profit], [Total Quantity] )", "$#,0.00"),
    ]

    model = {
        "name": PROJECT_NAME,
        "compatibilityLevel": 1550,
        "model": {
            "culture": "en-US",
            "tables": [
                {
                    "name": "Sales",
                    "columns": sales_columns,
                    "partitions": [
                        {
                            "name": "Sales",
                            "mode": "import",
                            "source": {
                                "type": "m",
                                "expression": m_csv_expression(),
                            },
                        }
                    ],
                    "measures": measures,
                },
            ],
        },
    }

    semantic_dir = PBI_DIR / f"{PROJECT_NAME}.SemanticModel"
    write_json(semantic_dir / "definition.pbism", {"version": "1.0"})
    write_json(semantic_dir / "model.bim", model)


def create_report_shell() -> None:
    def literal(text: str) -> dict:
        return {"expr": {"Literal": {"Value": f"'{text}'"}}}

    def measure_select(source: str, table: str, name: str) -> dict:
        return {
            "Measure": {"Expression": {"SourceRef": {"Source": source}}, "Property": name},
            "Name": f"{table}.{name}",
            "NativeReferenceName": name,
        }

    def column_select(source: str, table: str, name: str) -> dict:
        return {
            "Column": {"Expression": {"SourceRef": {"Source": source}}, "Property": name},
            "Name": f"{table}.{name}",
            "NativeReferenceName": name,
        }

    def query(selects: list[dict]) -> dict:
        from_items = [{"Name": "s", "Entity": "Sales", "Type": 0}]
        return {"Version": 2, "From": from_items, "Select": selects}

    def visual(
        visual_type: str,
        title: str,
        x: float,
        y: float,
        width: float,
        height: float,
        projections: dict,
        prototype_query: dict,
        tab_order: int,
    ) -> dict:
        config = {
            "name": str(uuid4()),
            "layouts": [
                {
                    "id": 0,
                    "position": {
                        "x": x,
                        "y": y,
                        "z": tab_order,
                        "width": width,
                        "height": height,
                        "tabOrder": tab_order,
                    },
                }
            ],
            "singleVisual": {
                "visualType": visual_type,
                "projections": projections,
                "prototypeQuery": prototype_query,
                "drillFilterOtherVisuals": True,
                "hasDefaultSort": True,
                "objects": {},
                "vcObjects": {
                    "title": [
                        {
                            "properties": {
                                "text": literal(title),
                                "show": {"expr": {"Literal": {"Value": "true"}}},
                            }
                        }
                    ]
                },
            },
        }
        return {
            "config": json.dumps(config),
            "query": json.dumps(prototype_query),
            "dataTransforms": "{}",
        }

    total_revenue = measure_select("s", "Sales", "Total Revenue")
    total_profit = measure_select("s", "Sales", "Total Profit")
    profit_margin = measure_select("s", "Sales", "Profit Margin")
    total_orders = measure_select("s", "Sales", "Total Orders")
    aov = measure_select("s", "Sales", "Avg Order Value")
    customer_count = measure_select("s", "Sales", "Customer Count")
    repeat_customer = measure_select("s", "Sales", "Repeat Customer")
    avg_customer_spend = measure_select("s", "Sales", "Avg Customer Spend")

    kpi_visuals = [
        visual("card", "Revenue", 20, 30, 210, 95, {"Values": [{"queryRef": "Sales.Total Revenue"}]}, query([total_revenue]), 0),
        visual("card", "Profit", 250, 30, 210, 95, {"Values": [{"queryRef": "Sales.Total Profit"}]}, query([total_profit]), 1),
        visual("card", "Margin", 480, 30, 210, 95, {"Values": [{"queryRef": "Sales.Profit Margin"}]}, query([profit_margin]), 2),
        visual("card", "Orders", 710, 30, 210, 95, {"Values": [{"queryRef": "Sales.Total Orders"}]}, query([total_orders]), 3),
        visual("card", "AOV", 940, 30, 210, 95, {"Values": [{"queryRef": "Sales.Avg Order Value"}]}, query([aov]), 4),
    ]

    executive_visuals = kpi_visuals + [
        visual(
            "lineChart",
            "Sales Trend",
            20,
            150,
            540,
            300,
            {"Category": [{"queryRef": "Sales.OrderDate", "active": True}], "Y": [{"queryRef": "Sales.Total Revenue"}]},
            query([column_select("s", "Sales", "OrderDate"), total_revenue]),
            5,
        ),
        visual(
            "barChart",
            "Revenue by Category",
            590,
            150,
            560,
            300,
            {"Category": [{"queryRef": "Sales.ProductCategory", "active": True}], "Y": [{"queryRef": "Sales.Total Revenue"}]},
            query([column_select("s", "Sales", "ProductCategory"), total_revenue]),
            6,
        ),
        visual(
            "treemap",
            "Profit by Region",
            20,
            475,
            540,
            235,
            {"Category": [{"queryRef": "Sales.Region", "active": True}], "Values": [{"queryRef": "Sales.Total Profit"}]},
            query([column_select("s", "Sales", "Region"), total_profit]),
            7,
        ),
        visual(
            "barChart",
            "Revenue by Segment",
            590,
            475,
            560,
            235,
            {"Category": [{"queryRef": "Sales.Segment", "active": True}], "Y": [{"queryRef": "Sales.Total Revenue"}]},
            query([column_select("s", "Sales", "Segment"), total_revenue]),
            8,
        ),
    ]

    product_visuals = [
        visual(
            "tableEx",
            "Top Products",
            20,
            30,
            650,
            650,
            {
                "Values": [
                    {"queryRef": "Sales.ProductName"},
                    {"queryRef": "Sales.Total Revenue"},
                    {"queryRef": "Sales.Total Profit"},
                    {"queryRef": "Sales.Total Quantity"},
                    {"queryRef": "Sales.Profit Margin"},
                ]
            },
            query(
                [
                    column_select("s", "Sales", "ProductName"),
                    total_revenue,
                    total_profit,
                    measure_select("s", "Sales", "Total Quantity"),
                    profit_margin,
                ]
            ),
            0,
        ),
        visual(
            "barChart",
            "Top 10 Products by Revenue",
            700,
            30,
            450,
            310,
            {"Category": [{"queryRef": "Sales.ProductName", "active": True}], "Y": [{"queryRef": "Sales.Total Revenue"}]},
            query([column_select("s", "Sales", "ProductName"), total_revenue]),
            1,
        ),
        visual(
            "scatterChart",
            "Revenue vs Margin",
            700,
            370,
            450,
            310,
            {
                "Category": [{"queryRef": "Sales.ProductName", "active": True}],
                "X": [{"queryRef": "Sales.Total Revenue"}],
                "Y": [{"queryRef": "Sales.Profit Margin"}],
                "Size": [{"queryRef": "Sales.Total Quantity"}],
            },
            query([column_select("s", "Sales", "ProductName"), total_revenue, profit_margin, measure_select("s", "Sales", "Total Quantity")]),
            2,
        ),
    ]

    customer_visuals = [
        visual("card", "Repeat Customer", 20, 30, 260, 100, {"Values": [{"queryRef": "Sales.Repeat Customer"}]}, query([repeat_customer]), 0),
        visual("card", "Customer Count", 310, 30, 260, 100, {"Values": [{"queryRef": "Sales.Customer Count"}]}, query([customer_count]), 1),
        visual("card", "Avg Customer Spend", 600, 30, 260, 100, {"Values": [{"queryRef": "Sales.Avg Customer Spend"}]}, query([avg_customer_spend]), 2),
        visual(
            "barChart",
            "Segment Analysis",
            20,
            165,
            560,
            300,
            {"Category": [{"queryRef": "Sales.Segment", "active": True}], "Y": [{"queryRef": "Sales.Total Revenue"}]},
            query([column_select("s", "Sales", "Segment"), total_revenue]),
            3,
        ),
        visual(
            "donutChart",
            "Revenue Share by Segment",
            620,
            165,
            520,
            300,
            {"Category": [{"queryRef": "Sales.Segment", "active": True}], "Y": [{"queryRef": "Sales.Total Revenue"}]},
            query([column_select("s", "Sales", "Segment"), total_revenue]),
            4,
        ),
    ]

    forecast_visuals = [
        visual(
            "lineChart",
            "Sales Forecast",
            20,
            30,
            1130,
            620,
            {"Category": [{"queryRef": "Sales.OrderDate", "active": True}], "Y": [{"queryRef": "Sales.Total Revenue"}]},
            query([column_select("s", "Sales", "OrderDate"), total_revenue]),
            0,
        )
    ]

    report_dir = PBI_DIR / f"{PROJECT_NAME}.Report"
    write_json(
        PBI_DIR / f"{PROJECT_NAME}.pbip",
        {
            "version": "1.0",
            "artifacts": [{"report": {"path": f"{PROJECT_NAME}.Report"}}],
            "settings": {"enableAutoRecovery": True},
        },
    )
    write_json(
        report_dir / "definition.pbir",
        {
            "$schema": "https://developer.microsoft.com/json-schemas/fabric/item/report/definitionProperties/1.0.0/schema.json",
            "version": "1.0",
            "datasetReference": {"byPath": {"path": f"../{PROJECT_NAME}.SemanticModel"}},
        },
    )
    write_json(
        report_dir / "report.json",
        {
            "config": json.dumps(
                {
                    "version": "5.58",
                    "themeCollection": {
                        "baseTheme": {"name": "CY24SU06", "version": "5.58", "type": 2},
                        "customTheme": None,
                    },
                }
            ),
            "layoutOptimization": 0,
            "sections": [
                {
                    "name": "ExecutiveOverview",
                    "displayName": "Executive Overview",
                    "width": 1280,
                    "height": 720,
                    "visualContainers": executive_visuals,
                },
                {
                    "name": "ProductPerformance",
                    "displayName": "Product Performance",
                    "width": 1280,
                    "height": 720,
                    "visualContainers": product_visuals,
                },
                {
                    "name": "CustomerAnalytics",
                    "displayName": "Customer Analytics",
                    "width": 1280,
                    "height": 720,
                    "visualContainers": customer_visuals,
                },
                {
                    "name": "SalesForecast",
                    "displayName": "Sales Forecast",
                    "width": 1280,
                    "height": 720,
                    "visualContainers": forecast_visuals,
                },
            ],
        },
    )
    (report_dir / ".pbi").mkdir(parents=True, exist_ok=True)
    write_json(report_dir / ".pbi" / "localSettings.json", {"version": "1.0"})


def create_pbids_file() -> None:
    write_json(
        PBI_DIR / "coffee_shop_sales_clean.pbids",
        {
            "version": "0.1",
            "connections": [
                {
                    "details": {
                        "protocol": "file",
                        "address": {"path": str(DATA_CSV)},
                    },
                    "mode": "Import",
                }
            ],
        },
    )


def add_table(ws, start_row: int, start_col: int, dataframe: pd.DataFrame, title: str) -> tuple[int, int]:
    ws.cell(start_row, start_col, title)
    ws.cell(start_row, start_col).font = Font(bold=True, color="FFFFFF")
    ws.cell(start_row, start_col).fill = PatternFill("solid", fgColor="0F2A43")
    for col_idx, col_name in enumerate(dataframe.columns, start_col):
        cell = ws.cell(start_row + 1, col_idx, col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F80ED")
    for row_idx, row in enumerate(dataframe.itertuples(index=False), start_row + 2):
        for col_idx, value in enumerate(row, start_col):
            ws.cell(row_idx, col_idx, value)
    return start_row + len(dataframe) + 1, start_col + len(dataframe.columns) - 1


def create_preview_workbook() -> None:
    df = pd.read_csv(DATA_CSV, parse_dates=["OrderDate"])
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Overview"
    ws.sheet_view.showGridLines = False

    dark = "0F2A43"
    blue = "2F80ED"
    gray = "E5E7EB"
    border = Border(bottom=Side(style="thin", color=gray))

    ws["A1"] = "Retail Sales & Revenue Dashboard"
    ws["A1"].font = Font(size=20, bold=True, color=dark)
    ws["A2"] = "Power BI project preview - Coffee Shop Sales"
    ws["A2"].font = Font(size=11, color="6B7280")

    kpis = [
        ("Revenue", df["Revenue"].sum(), "$#,##0"),
        ("Profit", df["Profit"].sum(), "$#,##0"),
        ("Margin", df["Profit"].sum() / df["Revenue"].sum(), "0.00%"),
        ("Orders", df["OrderID"].nunique(), "#,##0"),
        ("AOV", df["Revenue"].sum() / df["OrderID"].nunique(), "$#,##0.00"),
    ]
    for i, (label, value, fmt) in enumerate(kpis):
        col = 1 + i * 2
        ws.cell(4, col, label)
        ws.cell(4, col).font = Font(bold=True, color="6B7280")
        ws.cell(5, col, value)
        ws.cell(5, col).number_format = fmt
        ws.cell(5, col).font = Font(size=16, bold=True, color=dark)
        ws.cell(5, col).border = border

    trend = df.groupby(df["OrderDate"].dt.to_period("M")).agg(Revenue=("Revenue", "sum"), Profit=("Profit", "sum")).reset_index()
    trend["OrderDate"] = trend["OrderDate"].astype(str)
    add_table(ws, 8, 1, trend, "Monthly Sales Trend")

    category = df.groupby("ProductCategory", as_index=False).agg(Revenue=("Revenue", "sum"), Profit=("Profit", "sum")).sort_values("Revenue", ascending=False)
    add_table(ws, 8, 5, category, "Revenue by Category")

    region = df.groupby("Region", as_index=False).agg(Revenue=("Revenue", "sum"), Profit=("Profit", "sum")).sort_values("Profit", ascending=False)
    add_table(ws, 22, 1, region, "Profit by Region")

    top_products = (
        df.groupby("ProductName", as_index=False)
        .agg(Revenue=("Revenue", "sum"), Profit=("Profit", "sum"), Quantity=("Quantity", "sum"))
        .sort_values("Revenue", ascending=False)
        .head(10)
    )
    add_table(ws, 22, 5, top_products, "Top 10 Products")

    line = LineChart()
    line.title = "Sales Trend"
    line.y_axis.title = "Revenue"
    line.x_axis.title = "Month"
    line.add_data(Reference(ws, min_col=2, min_row=9, max_row=9 + len(trend)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=10, max_row=9 + len(trend)))
    line.height = 7
    line.width = 14
    ws.add_chart(line, "A36")

    bar = BarChart()
    bar.title = "Revenue by Category"
    bar.y_axis.title = "Revenue"
    bar.x_axis.title = "Category"
    bar.add_data(Reference(ws, min_col=6, min_row=9, max_row=9 + len(category)), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=5, min_row=10, max_row=9 + len(category)))
    bar.height = 7
    bar.width = 16
    ws.add_chart(bar, "I36")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    for idx in range(1, 14):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    output = PBI_DIR / f"{PROJECT_NAME} Preview.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    if not DATA_CSV.exists():
        raise FileNotFoundError(f"Missing processed data: {DATA_CSV}")
    create_model_bim()
    create_report_shell()
    create_pbids_file()
    create_preview_workbook()
    print(f"Created Power BI project: {PBI_DIR / (PROJECT_NAME + '.pbip')}")
    print(f"Created Power BI data source: {PBI_DIR / 'coffee_shop_sales_clean.pbids'}")
    print(f"Created preview workbook: {PBI_DIR / (PROJECT_NAME + ' Preview.xlsx')}")


if __name__ == "__main__":
    main()
